from tokenize import String
import openpyxl as xl 
from openpyxl.worksheet.table import Table, TableStyleInfo

# Getting the Excel's Workbook with openpyxl
filename = r"C:\Users\nicor\OneDrive - UNIVERSIDAD INDUSTRIAL DE SANTANDER\Desktop\Financial\NeoCrypto.xlsx" # you gotta edit this line according to the location of your ATD
archivo = xl.load_workbook(filename)

def modifyCellValue(row: int, column: int, sheet, value, format = "General"):
    hoja = archivo[sheet]
    hoja.cell(row, column).value = value
    hoja.cell(row, column).number_format = format

def saveWb():
    archivo.save(filename)

def modifyTable(cell1, cell2, sheet, name):
    hoja = archivo[sheet]
    column1= cell1[0]
    row1= cell1[1]
    column2= cell2[0]
    row2= cell2[1]
    hoja.tables[name].ref = "{}{}:{}{}".format(column1, row1, column2, row2)
